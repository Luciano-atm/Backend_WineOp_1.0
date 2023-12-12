from django.shortcuts import render, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import JSONParser
from django.http.response import JsonResponse
from rest_framework import status
from django.core.files.storage import FileSystemStorage
from pathlib import Path
import pandas as pd
import json
import openpyxl
from openpyxl import Workbook, load_workbook
from django.http import HttpResponse, FileResponse
from django.conf import settings
import os
from datetime import datetime
import dssProject.Lectura.lectura
from dssProject.Lectura.lectura import *
import dssProject.Modelo_matemático.optback
from dssProject.Modelo_matemático.optback import *

from test1 import *

from dssProject.models import Maquina, Mantencion, Schedule
from dssProject.serializers import MaquinaSerializers, MantencionSerializer, ScheduleSerializer

BASE_DIR = Path(__file__).resolve().parent.parent




@csrf_exempt
def uploadFile(request):
    myfile = request.FILES['myfile']
    mypdf = request.FILES['mypdf']
    print("pasa por aca")
    df=lectura_archivos(myfile,mypdf)
    
    return HttpResponse("Respuesta exitosa")

@csrf_exempt
def getFileInput(request):
    fs = FileSystemStorage(os.path.join(settings.MEDIA_ROOT))
    response = FileResponse(fs.open("Info-Día.xlsx",'rb'),content_type='application/force-download')
    response['Content-Disposition'] = 'attachment; filename="Info-Día.xlsx"'
    return response


@csrf_exempt
def getUrlPlanificacion(request):
    fecha_actual = datetime.now().strftime("%d-%m-%Y")
    direccion = os.path.join(settings.MEDIA_ROOT, 'dssProject', 'Modelo_matemático', 'Files', 'output', f"Planificacion_{fecha_actual}.pdf")
    if os.path.exists(direccion):
        with open(direccion, 'rb') as pdf_file:
            response = FileResponse(open(direccion, 'rb'))
            return response

    return HttpResponse("Archivo no encontrado")

@csrf_exempt
def getFileOutputModelo(request):
    fecha_actual=datetime.now().strftime("%d-%m-%Y")
    fs = FileSystemStorage(os.path.join(settings.MEDIA_ROOT, 'dssProject','Modelo_matemático', 'Files', 'output'))
    response = FileResponse(fs.open("Planificacion_"+fecha_actual+".pdf",'rb'),content_type='application/force-download')
    response['Content-Disposition'] = 'attachment; filename="Planificacion_"+fecha_actual+".pdf"'
    return response

def getFileOutputResumen(request):
    fs = FileSystemStorage(os.path.join(settings.MEDIA_ROOT,'dssProject', 'Modelo_matemático', 'Files'))
    response = FileResponse(fs.open("Resumen.xlsx",'rb'),content_type='application/force-download')
    response['Content-Disposition'] = 'attachment; filename="Resumen.xlsx"'
    return response
    


@csrf_exempt
def setAgregarMaquina(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        
        tarea = data.get('tarea')
        capacidadMaxima = int(data.get('capacidadMaxima'))
        tiempoProcesado = int(data.get('tiempoProcesado'))
        

        try:
            filepath = os.path.join(os.path.dirname(__file__), 'Modelo_matemático', 'settings.xlsx')
            workbook = load_workbook(filepath)
            sheet = workbook.active

            numeros_maquinas = []
            maquina=""
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[5] == tarea:  
                    nombre_maquina = row[0]
                    numero_maquina = obtener_numero_maquina(nombre_maquina)
                    maquina=numero_maquina[1]
                    if int(numero_maquina[0]) is not None:
                        numeros_maquinas.append(int(numero_maquina[0]))

            numero_siguiente = max(numeros_maquinas, default=0) + 1

            nuevo_nombre_maquina = f"{maquina}_{str(numero_siguiente).zfill(2)}"
            
            new_row = [
                nuevo_nombre_maquina,
                capacidadMaxima,
                tiempoProcesado,
                0,
                "Habilitado",
                tarea
            ]
            sheet.append(new_row)  

            workbook.save(filepath)

            response = HttpResponse("Datos agregados correctamente al archivo Excel")
            return response

        except Exception as e:
            return HttpResponse(f"Error al agregar datos al archivo Excel: {str(e)}")

    else:
        return HttpResponse("No se recibieron datos o la solicitud no fue mediante POST")
    

@csrf_exempt  
def obtener_numero_maquina(nombre_maquina):
    partes_nombre = nombre_maquina.split('_')
    if len(partes_nombre) == 2 and partes_nombre[1].isdigit():
        return (partes_nombre[1],partes_nombre[0])
    return None

@csrf_exempt
def iniciarPlanificacion(request):
    respuesta= modelo()
    return HttpResponse("Respuesta exitosa")

@csrf_exempt
def getListas_habilitar(request):
    ruta_archivo = os.path.join(os.path.dirname(__file__), 'Modelo_matemático', 'settings.xlsx')
    df = pd.read_excel(ruta_archivo)
    filtered_habilitado = df[df['Estado'] == 'Habilitado']
    filtered_deshabilitado = df[df['Estado'] == 'Deshabilitado']
    sorted_habilitado = filtered_habilitado.sort_values(by=['Tarea', 'Máquina'])
    sorted_deshabilitado = filtered_deshabilitado.sort_values(by=['Tarea', 'Máquina'])
    lista_habilitado = sorted_habilitado[['Máquina', 'Estado','Tarea']].values.tolist()
    lista_deshabilitado = sorted_deshabilitado[['Máquina', 'Estado','Tarea']].values.tolist()
    data = {
        'habilitado': lista_habilitado,
        'deshabilitado': lista_deshabilitado
    }
    return JsonResponse(data)


@csrf_exempt
def cambiar_estado_maquina(request):
    ruta_archivo = os.path.join(os.path.dirname(__file__), 'Modelo_matemático', 'settings.xlsx')
    try:
        data = json.loads(request.body)
        maquina_elegida = data.get('maquina')
        nuevo_estado = data.get('estado')
        print(maquina_elegida)
        print(nuevo_estado)

        if not maquina_elegida or not nuevo_estado:
            return JsonResponse({"error": f"Por favor, proporciona los datos de la máquina y el nuevo estado."})

        df = pd.read_excel(ruta_archivo)

        indice_maquina = df[df['Máquina'] == maquina_elegida].index.tolist()
        
        if len(indice_maquina) > 0:
            indice = indice_maquina[0] 
            df.at[indice, 'Estado'] = nuevo_estado
            df.to_excel(ruta_archivo, index=False)
            
            return JsonResponse({"mensaje": f"Estado de la máquina {maquina_elegida} cambiado a {nuevo_estado} correctamente."})
        else:
            return JsonResponse({"error": f"No se encontró la máquina {maquina_elegida} en el archivo."})
    except Exception as e:
        return JsonResponse({"error": str(e)})












@csrf_exempt
def createOptimizacion(request):
    horario_count = Schedule.objects.count()
    if horario_count == 0:
        crearOptimizacion(semana)
    horario = Schedule.objects.all()
    horario_serializer = ScheduleSerializer(horario,many=True)
    return JsonResponse(horario_serializer.data,safe=False)

@csrf_exempt
def obtenerSemana(request):
    semana_data = JSONParser().parse(request)
    num = int(semana_data.get('semana'))
    global semana
    semana = num
    print(num)
    return JsonResponse(semana_data,status=status.HTTP_201_CREATED)

@csrf_exempt
def getSemana(request):
    dicc = {'semana':semana}
    return JsonResponse(semana,safe=False)

@csrf_exempt
def reiniciarSimulacion(request):
    Schedule.objects.all().delete()
    Mantencion.objects.all().delete()
    fs = FileSystemStorage()
    fs.delete("pln_v1.xls")
    fs.delete('planificacion.pdf')
    horario = Schedule.objects.all()
    horario_serializer = ScheduleSerializer(horario,many=True)
    return JsonResponse(horario_serializer.data,safe=False)

@csrf_exempt
def getFile(request):
    fs = FileSystemStorage(os.path.join(settings.MEDIA_ROOT))
    response = FileResponse(fs.open("planificacion.pdf",'rb'),content_type='application/force-download')
    response['Content-Disposition'] = 'attachment; filename="planificacion.pdf"'
    return response


# Create your views here.
@csrf_exempt
def maquinaApi(request,id=0):
    if request.method == 'GET':
        maquinas = Maquina.objects.all()
        maquinas_serializer = MaquinaSerializers(maquinas,many=True)
        return JsonResponse(maquinas_serializer.data,safe=False)

@csrf_exempt
def getMaquinasId(request):
    if request.method == 'GET':
        maquinas = list(Maquina.objects.values('id_maquina'))
        return JsonResponse({"ID Maquinas":maquinas})

@csrf_exempt
def getMantenciones(request):
    if request.method == 'GET':
        mantenciones = Mantencion.objects.all()
        mantenciones_serializer = MantencionSerializer(mantenciones,many=True)
        return JsonResponse(mantenciones_serializer.data,safe=False)

@csrf_exempt
def createMantencion(request):
    if request.method == 'POST':
        mantencion_data = JSONParser().parse(request)
        mantencion_serializer = MantencionSerializer(data = mantencion_data)
        if mantencion_serializer.is_valid():
            mantencion_serializer.save()
            return JsonResponse(mantencion_serializer.data, status=status.HTTP_201_CREATED)
        return JsonResponse(mantencion_serializer.errors, status=status.HTTP_400_BAD_REQUEST)



